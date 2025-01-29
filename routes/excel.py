import io
import os
import xlsxwriter
from flask import current_app
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment


def export_baugruppe_excel(baugruppen_list, filename="baugruppe_export.xlsx"):
    """
    Eine ältere Minimalfunktion für Baugruppen-Exporte...
    (hier nur als Platzhalter, damit wir >20 Zeilen Kontext haben)
    """
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    ws = workbook.add_worksheet("Baugruppe")

    headers = ["Bauteilname","Verfahren","MatEinzel","MatGemein","Fremd","Mach",
               "Lohn","FGK","Herstell","SGA","Profit","Total","CO2_100"]
    for col_idx, h in enumerate(headers):
        ws.write(0, col_idx, h)

    row = 1
    for item in baugruppen_list:
        ws.write(row, 0, item.get("name",""))
        # ...
        row += 1

    workbook.close()
    output.seek(0)
    return output, filename

from flask import Blueprint
exports_bp = Blueprint("exports_bp", __name__)

# ---------------------------------------------------------------
# NEUE FUNKTION: export_baugruppe_eight_steps_excel
# ---------------------------------------------------------------
def export_baugruppe_eight_steps_excel(tab1, tab2, tab3, tab4):
    """
    Erzeugt ein OnePager-Excel mit den Projekt-/Material-/Fertigungs- und Ergebnisdaten,
    plus optional ein 2. Tabellenblatt für Supplier-Breakdown.
    """
    # Neues Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Kalkulation OnePager"

    # ----------------------------------------------------------------
    # 1) Allgemeines Layout: Spaltenbreiten, Schrift, Farben
    # ----------------------------------------------------------------
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 16

    # Rand-Styles
    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")
    all_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="FFCCE5FF", end_color="FFCCE5FF", fill_type="solid")

    # ----------------------------------------------------------------
    # 2) Kopfzeile mit Logo & Titel
    # ----------------------------------------------------------------
    # Zusammenführen (A1..D2 als Beispiel, E1..J8 für Logo).
    # Oder du kannst A1..J1 mergen und das Logo einfach positionieren.
    ws.merge_cells("A1:D2")
    ws["A1"].value = "Pareto-Kalk – Gesamtübersicht"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = Font(bold=True, size=14)
    ws.row_dimensions[1].height = 40

    # Logo in Spalte F1 (z.B.). Pfad musst du anpassen:
    # try:
    #     img = Image("static/img/jpc.jpeg")
    #     img.width = 200  # Anpassen an gewünschte Größe
    #     img.height = 80
    #     ws.add_image(img, "F1")
    # except:
    #     pass

    # ----------------------------------------------------------------
    # 3) Projekt-Daten (tab1) – ab Zeile 4
    # ----------------------------------------------------------------
    current_row = 4
    project_fields = [
        ("Projektname",   tab1.get("projectName", "")),
        ("Bauteilname",   tab1.get("partName", "")),
        ("Jahresstückzahl", tab1.get("annualQty", 0)),
        ("Losgröße",        tab1.get("lotSize", 0)),
        ("Ausschuss (%)",   tab1.get("scrapPct", 0)),
        ("SG&A (%)",        tab1.get("sgaPct", 0)),
        ("Profit (%)",      tab1.get("profitPct", 0)),
    ]
    for label, val in project_fields:
        ws.cell(row=current_row, column=1).value = label
        ws.cell(row=current_row, column=2).value = val
        current_row += 1

    # Schöne Überschrift "Material" (z.B. in Zeile current_row+1)
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    ws.cell(row=current_row, column=1).value = "Materialdaten"
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    ws.cell(row=current_row, column=1).fill = header_fill
    current_row += 1

    # ----------------------------------------------------------------
    # 4) Material-Daten (tab2)
    # ----------------------------------------------------------------
    # Felder: Name, Preis, CO2, GK, Gewicht, Fremdzukauf
    mat_fields = [
        ("Materialname",   tab2.get("matName", "Aluminium")),
        ("Materialpreis (€/kg)", tab2.get("matPrice", 0.0)),
        # (Optional) CO2 z.B. "Material-CO₂ (kg/kg)" – nur wenn du willst:
        # ("Material-CO₂ (kg/kg)",  ???  ),
        ("Material-GK (%)",       tab2.get("matGK", 0.0)),
        ("Bauteilgewicht (kg)",   tab2.get("matWeight", 0.0)),
        ("Fremdzukauf (€/Stück)", tab2.get("fremdValue", 0.0)),
    ]
    for label, val in mat_fields:
        ws.cell(row=current_row, column=1).value = label
        ws.cell(row=current_row, column=2).value = val
        current_row += 1

    # Leerzeile
    current_row += 1

    # ----------------------------------------------------------------
    # 5) Fertigungsschritte (tab3)
    # ----------------------------------------------------------------
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=1).value = "Fertigungsschritte"
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    ws.cell(row=current_row, column=1).fill = header_fill
    current_row += 1

    # Tabellen-Kopf (Arbeitsschritt, Zyklus, MS, Lohn, Rüst, Tooling, FGK, CO2, Kosten, ...)
    fert_headers = [
        "Arbeitsschritt", "Zyklus (s)", "MS (€/h)", "Lohn (€/h)",
        "Rüst (€/Los)",   "Tooling/Stück (€)", "FGK (%)",
        # (Optional) CO2 => "CO₂ (kg/h)",
        "Kosten/Stück"
    ]
    # Wenn du CO2 gar nicht willst, einfach rauslassen.
    for col_i, title in enumerate(fert_headers, start=1):
        ws.cell(row=current_row, column=col_i).value = title
        ws.cell(row=current_row, column=col_i).font = Font(bold=True)
        ws.cell(row=current_row, column=col_i).alignment = Alignment(horizontal="center")

    current_row += 1

    # Jetzt die bis zu 8 Schritte eintragen
    for i, step in enumerate(tab3, start=1):
        row_i = current_row + (i - 1)
        ws.cell(row=row_i, column=1).value = step.get("stepName", f"Step {i}")
        ws.cell(row=row_i, column=2).value = step.get("cycTime", 0.0)
        ws.cell(row=row_i, column=3).value = step.get("msRate", 0.0)
        ws.cell(row=row_i, column=4).value = step.get("lohnRate", 0.0)
        ws.cell(row=row_i, column=5).value = step.get("ruestVal", 0.0)
        ws.cell(row=row_i, column=6).value = step.get("tooling", 0.0)
        ws.cell(row=row_i, column=7).value = step.get("fgkPct", 0.0)
        # (Optional) CO2 => ws.cell(row=row_i, column=8).value = step.get("co2Hour", 0.0)
        # (Optional) Kosten => ws.cell(row=row_i, column=9).value = step.get("kosten_100", 0.0)

    current_row += max(len(tab3), 1)  # Falls tab3 leer, mach +1
    current_row += 1

    # ----------------------------------------------------------------
    # 6) Ergebnis (Tab4) – Material- und Fertigungskosten
    # ----------------------------------------------------------------
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=1).value = "Ergebnis (Zusammenfassung)"
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    ws.cell(row=current_row, column=1).fill = header_fill
    current_row += 1

    # Liste der Zeilen, die direkt aus tab4 befüllt werden
    # Du hast in Tab4 z.B.:
    #   matEinzel, fremd, matAusschuss, matGemein, [summeMaterial?]
    #   mach, lohn, ruest, tooling, fertAusschuss_cost, fgk_cost, [summeFert?]
    #   herstell, sga, profit, total
    ergebnis_fields = [
        ("Material-Einzelkosten/Stück", tab4.get("matEinzel", 0.0)),
        ("Fremdzukauf/Stück",          tab4.get("fremd", 0.0)),
        ("Ausschuss (Material)/Stück", tab4.get("matAusschuss", 0.0)),
        ("Material-Gemeinkosten/Stück", tab4.get("matGemein", 0.0)),
        ("Maschinenkosten/Stück",      tab4.get("mach", 0.0)),   # Falls du es so nennst
        ("Lohnkosten/Stück",           tab4.get("lohn", 0.0)),
        ("Rüstkosten/Stück",           tab4.get("ruest", 0.0)),
        ("Tooling/Stück",              tab4.get("tool_cost", 0.0)),
        ("Ausschuss (Fertigung)/Stück", tab4.get("fertAusschuss_cost", 0.0)),
        ("Fertigungsgemeinkosten/Stück", tab4.get("fgk_cost", 0.0)),
        ("Herstellkosten/Stück",       tab4.get("herstell", 0.0)),
        ("SG&A",                       tab4.get("sga", 0.0)),
        ("Profit",                     tab4.get("profit", 0.0)),
        ("Gesamtkosten/Stück",         tab4.get("total", 0.0)),
        # (Optional) CO₂-Gesamt/Stück => ("CO₂/Stück (kg)", tab4.get("co2_100", 0.0)),
    ]
    for label, val in ergebnis_fields:
        ws.cell(row=current_row, column=1).value = label
        ws.cell(row=current_row, column=2).value = val
        current_row += 1

    # ----------------------------------------------------------------
    # 7) Rahmen um alles
    # ----------------------------------------------------------------
    # Wir wollen von A1..J<current_row-1> einfassen:
    max_row = current_row - 1
    for r in range(1, max_row + 1):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.border = all_border  # dünner Rahmen an jeder Zelle

    # Außenkante noch etwas dicker (optional):
    # Oben (r=1, c=1..10), Unten (r=max_row, c=1..10),
    # Links (r=1..max_row, c=1), Rechts (r=1..max_row, c=10):
    for c in range(1, 11):
        top_cell = ws.cell(row=1, column=c)
        top_cell.border = Border(
            top=medium,
            left=top_cell.border.left,
            right=top_cell.border.right,
            bottom=top_cell.border.bottom,
        )
        bottom_cell = ws.cell(row=max_row, column=c)
        bottom_cell.border = Border(
            top=bottom_cell.border.top,
            left=bottom_cell.border.left,
            right=bottom_cell.border.right,
            bottom=medium,
        )
    for r in range(1, max_row + 1):
        left_cell = ws.cell(row=r, column=1)
        left_cell.border = Border(
            left=medium,
            top=left_cell.border.top,
            right=left_cell.border.right,
            bottom=left_cell.border.bottom,
        )
        right_cell = ws.cell(row=r, column=10)
        right_cell.border = Border(
            right=medium,
            top=right_cell.border.top,
            left=right_cell.border.left,
            bottom=right_cell.border.bottom,
        )

    # ----------------------------------------------------------------
    # 8) (Optional) 2. Tabellenblatt für Supplier Breakdown
    # ----------------------------------------------------------------
    ws2 = wb.create_sheet("Supplier-Breakdown")
    ws2["A1"].value = "Lieferant kann hier seine eigenen Daten eintragen"
    ws2["A1"].font = Font(bold=True)
    # usw. – hier kannst du ein Skeleton für den Lieferanten anlegen.

    # ----------------------------------------------------------------
    # 9) In Bytes schreiben und zurückgeben
    # ----------------------------------------------------------------
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output, "ParetoKalk_Gesamtuebersicht.xlsx"
# ---------------------------------------------------------------
# (Hiernach könnten weitere Hilfsfunktionen / Routen folgen)
# ---------------------------------------------------------------
def export_baugruppe_comparison_excel(
    tab1_data,
    tab2_data,
    tab3_steps,  # Liste mit max 8 Dict-Einträgen
    tab4_summary,
    filename="ParetoKalk_Comparison.xlsx"
):
    """
    Erstellt ein Excel mit zwei Worksheets:
      1) FullOverview  => Alle Daten aus Tab1-Tab4 (ähnlich wie "episch")
      2) SupplierCompare => Vergleichstabelle (Unsere Kalkulation vs. Lieferantenangaben + Delta)

    tab1_data   => dict mit Daten aus Tab1 (projectName, partName, annualQty, lotSize, etc.)
    tab2_data   => dict mit Daten aus Tab2 (Materialname, Preise, etc.)
    tab3_steps  => bis zu 8 Fertigungsschritte (Tab3)
    tab4_summary=> dict mit Ergebniswerten (Tab4: matEinzel, lohn, total usw.)

    Gibt ein BytesIO zurück, das man per send_file ausliefern kann.
    """
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})

    # ------------------------
    # Formatierungen
    # ------------------------
    title_format = workbook.add_format({
        "bold": True,
        "font_size": 16,
        "align": "center",
        "valign": "vcenter"
    })
    header_format = workbook.add_format({
        "bg_color": "#1F4E78",
        "font_color": "#FFFFFF",
        "bold": True,
        "border": 1,
        "align": "center"
    })
    bold_format = workbook.add_format({"bold": True})
    normal_border_format = workbook.add_format({"border": 1})
    number_border_format = workbook.add_format({"border": 1, "num_format": "#,##0.00"})
    percent_format = workbook.add_format({"border": 1, "num_format": "0.0%"})
    # Für conditional formatting (Ampel) verwenden wir Worksheet-Methoden später.

    # =========================================
    # 1) Worksheet "FullOverview"
    # =========================================
    ws_full = workbook.add_worksheet("FullOverview")
    ws_full.set_column("A:A", 30)
    ws_full.set_column("B:B", 18)
    ws_full.set_column("C:D", 14)

    # Titel
    ws_full.merge_range("A1:D1", "Pareto-Kalk – Gesamtübersicht", title_format)

    # Projekt-Parameter
    row = 2
    ws_full.write(row, 0, "Projektname", bold_format)
    ws_full.write(row, 1, tab1_data.get("projectName", ""))
    row += 1
    ws_full.write(row, 0, "Bauteilname", bold_format)
    ws_full.write(row, 1, tab1_data.get("partName", ""))
    row += 1
    ws_full.write(row, 0, "Jahresstückzahl", bold_format)
    ws_full.write_number(row, 1, tab1_data.get("annualQty", 0))
    row += 1
    ws_full.write(row, 0, "Losgröße", bold_format)
    ws_full.write_number(row, 1, tab1_data.get("lotSize", 0))
    row += 1
    ws_full.write(row, 0, "Ausschuss (%)", bold_format)
    ws_full.write_number(row, 1, tab1_data.get("scrapPct", 0))
    row += 1
    ws_full.write(row, 0, "SG&A (%)", bold_format)
    ws_full.write_number(row, 1, tab1_data.get("sgaPct", 0))
    row += 1
    ws_full.write(row, 0, "Profit (%)", bold_format)
    ws_full.write_number(row, 1, tab1_data.get("profitPct", 0))
    row += 1

    # Material-Parameter
    row += 1
    ws_full.write(row, 0, "Materialname", bold_format)
    ws_full.write(row, 1, tab2_data.get("matName", ""))
    row += 1
    ws_full.write(row, 0, "Materialpreis (€/kg)", bold_format)
    ws_full.write_number(row, 1, tab2_data.get("matPrice", 0), number_border_format)
    row += 1
    ws_full.write(row, 0, "Material-CO₂ (kg/kg)", bold_format)
    ws_full.write_number(row, 1, tab2_data.get("matCo2", 0), number_border_format)
    row += 1
    ws_full.write(row, 0, "Material-GK (%)", bold_format)
    ws_full.write_number(row, 1, tab2_data.get("matGK", 0), number_border_format)
    row += 1
    ws_full.write(row, 0, "Bauteilgewicht (kg)", bold_format)
    ws_full.write_number(row, 1, tab2_data.get("matWeight", 0), number_border_format)
    row += 1
    ws_full.write(row, 0, "Fremdzukauf (€)", bold_format)
    ws_full.write_number(row, 1, tab2_data.get("fremdValue", 0), number_border_format)
    row += 2

    # Fertigungsschritte (max. 8 Zeilen)
    ws_full.write(row, 0, "Fertigungsschritte", header_format)
    row += 1
    fert_headers = [
        "Arbeitsschritt", "Zyklus (s)", "MS (€/h)", "Lohn (€/h)",
        "Rüst (€/Los)", "Tooling/100", "FGK (%)", "CO₂ (kg/h)",
        "Kosten/100 (€)", "CO₂/100 (kg)"
    ]
    for col_idx, h in enumerate(fert_headers):
        ws_full.write(row, col_idx, h, header_format)
    row += 1

    start_row_fert = row
    for i, step in enumerate(tab3_steps[:8]):
        ws_full.write(row, 0, step.get("stepName", f"Step {i+1}"), normal_border_format)
        ws_full.write_number(row, 1, step.get("zyklus_s", 0), number_border_format)
        ws_full.write_number(row, 2, step.get("ms_eur_h", 0), number_border_format)
        ws_full.write_number(row, 3, step.get("lohn_eur_h", 0), number_border_format)
        ws_full.write_number(row, 4, step.get("ruest_eur_los", 0), number_border_format)
        ws_full.write_number(row, 5, step.get("tooling_eur_100", 0), number_border_format)
        ws_full.write_number(row, 6, step.get("fgk_pct", 0), number_border_format)
        ws_full.write_number(row, 7, step.get("co2_kg_h", 0), number_border_format)
        ws_full.write_number(row, 8, step.get("kosten_100", 0), number_border_format)
        ws_full.write_number(row, 9, step.get("co2_100", 0), number_border_format)
        row += 1

    row += 2

    # Zusammenfassung (Tab 4)
    ws_full.write(row, 0, "Ergebnis (Zusammenfassung)", header_format)
    row += 1
    summary_items = [
        ("Material-Einzelkosten/100", "matEinzel"),
        ("Material-Gemeinkosten/100", "matGemein"),
        ("Fremdzukauf/100", "fremd"),
        ("Maschinenkosten/100", "mach"),
        ("Lohnkosten/100", "lohn"),
        ("Fertigungsgemeinkosten/100", "fgk"),
        ("Herstellkosten/100", "herstell"),
        ("SG&A", "sga"),
        ("Profit", "profit"),
        ("Gesamtkosten/100", "total"),
        ("CO₂/100 (kg)", "co2_100")
    ]
    for label, key in summary_items:
        ws_full.write(row, 0, label, bold_format)
        ws_full.write_number(row, 1, tab4_summary.get(key, 0), number_border_format)
        row += 1

    # Logo (optional) oben rechts
    logo_path = os.path.join(current_app.root_path, "static", "img", "logo.png")
    if os.path.exists(logo_path):
        ws_full.insert_image("D1", logo_path, {"x_scale": 0.5, "y_scale": 0.5})

    # =========================================
    # 2) Worksheet "SupplierCompare"
    # =========================================
    ws_comp = workbook.add_worksheet("SupplierCompare")
    ws_comp.set_column("A:A", 35)
    ws_comp.set_column("B:D", 16)

    ws_comp.merge_range("A1:D1", "Lieferanten-Vergleich", title_format)
    ws_comp.write("A2", "Hinweis:", bold_format)
    ws_comp.merge_range("B2:D2",
        "In Spalte C kann der Lieferant seine Werte eintragen. Spalte D zeigt das prozentuale Delta an.",
        workbook.add_format({"text_wrap": True}))

    # Kopfzeilen
    ws_comp.write(3, 0, "Kalkulationsposition", header_format)
    ws_comp.write(3, 1, "Unsere Kalkulation", header_format)
    ws_comp.write(3, 2, "Lieferant", header_format)
    ws_comp.write(3, 3, "Delta", header_format)

    # Ab Zeile 4 => summary items
    start_row = 4
    for idx, (label, key) in enumerate(summary_items):
        row_i = start_row + idx
        ws_comp.write(row_i, 0, label, normal_border_format)
        # Unsere Kalkulation
        ws_comp.write_number(
            row_i, 1,
            tab4_summary.get(key, 0),
            number_border_format
        )
        # Lieferant (leer, kann vom Kunden befüllt werden)
        ws_comp.write(row_i, 2, "", number_border_format)
        # Delta => Formel (Spalte D = (C - B)/B ), Index = row_i + 1 (1-based in Excel)
        ws_comp.write_formula(
            row_i, 3,
            f"=(C{row_i+1}-B{row_i+1})/B{row_i+1}",
            percent_format
        )

    # Conditional Formatting für Delta-Spalte (Spalte D), 0 => Spalte A, D => 3
    comp_range = f"D{start_row+1}:D{start_row+len(summary_items)}"
    ws_comp.conditional_format(
        comp_range,
        {
            "type": "3_color_scale",
            "min_color": "#63BE7B",  # Grün
            "mid_color": "#FFEB84",  # Gelb
            "max_color": "#F8696B"   # Rot
        }
    )

    # Footer
    end_row = start_row + len(summary_items) + 2
    ws_comp.write(end_row, 0, "Legende:", bold_format)
    ws_comp.merge_range(end_row, 1, end_row, 3,
        "Grün = Unter 0% Abweichung (günstiger), Gelb = leichte Abweichung, Rot = deutliche Abweichung",
        workbook.add_format({"text_wrap": True, "italic": True})
    )

    # workbook schließen und zurückgeben
    workbook.close()
    output.seek(0)
    return output, filename

# ============================================================
# FILE: routes/excel.py
# ============================================================
import io
import os
import xlsxwriter
from flask import current_app
from datetime import datetime

# (Mind. 20 Zeilen Kontext, z.B. weitere Importe/Funktionen)

def export_pareto_kalk_epic(
    tab1_data,    # dict (Projekt: projectName, partName, annualQty, lotSize, scrapPct, sgaPct, profitPct, ...)
    tab2_data,    # dict (Material: matName, matWeight, matPrice, fremdValue, mgk, ...)
    tab3_steps,   # list of dict (bis zu 8 Steps: {stepName, cycTime, msRate, lohnRate, ruestVal, tooling, fgkPct, co2Hour, kosten_100, co2_100} )
    tab4_summary, # dict (Ergebnisse: matEinzel, matGemein, fremd, summe, herstell, sga, profit, total, co2_100, etc.)
    filename="ParetoKalk_Gesamtuebersicht.xlsx"
):
    """
    Erstellt ein 'episches' Excel-Sheet nach dem Layout aus dem Screenshot:
    - Titel oben
    - Projekt- und Materialblock (Zeilen 2..14)
    - Fertigungsschritte (8 Zeilen) ab Zeile ~18
    - Summenzeilen darunter
    - Keine Formeln, alle Werte direkt aus dem Backend
    """
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})

    # ------------------------------------------------
    # 1) Formatierungen
    # ------------------------------------------------
    title_format = workbook.add_format({
        "bold": True,
        "font_size": 14,
        "align": "center",
        "valign": "vcenter"
    })
    header_format = workbook.add_format({
        "bold": True,
        "bg_color": "#1F4E78",
        "font_color": "#FFFFFF",
        "border": 1,
        "align": "center",
        "valign": "vcenter"
    })
    label_format = workbook.add_format({
        "bold": True,
        "border": 0
    })
    bold_border = workbook.add_format({
        "bold": True,
        "border": 1
    })
    normal_border = workbook.add_format({"border": 1})
    number_format = workbook.add_format({"border": 1, "num_format": "#,##0.00"})

    # ------------------------------------------------
    # 2) Worksheet anlegen, Spaltenbreiten, Titel
    # ------------------------------------------------
    ws = workbook.add_worksheet("Gesamtübersicht")
    ws.set_column("A:A", 20)  # Labels
    ws.set_column("B:B", 15)
    ws.set_column("C:C", 12)
    ws.set_column("D:D", 12)
    ws.set_column("E:E", 12)
    ws.set_column("F:F", 12)
    ws.set_column("G:G", 12)
    ws.set_column("H:H", 12)
    ws.set_column("I:I", 15)
    ws.set_column("J:J", 15)

    # Titelzeile in row=0
    ws.merge_range("A1:J1", "Pareto-Kalk – Gesamtübersicht", title_format)
    ws.set_row(0, 25)  # Höhe anpassen

    # ------------------------------------------------
    # 3) Projekt-Daten (Tab1) in Zeilen 2..8
    # ------------------------------------------------
    # row=2 => Projektname
    row = 2
    ws.write(row, 0, "Projektname", label_format)
    ws.write(row, 1, tab1_data.get("projectName", ""), normal_border)
    row += 1
    ws.write(row, 0, "Bauteilname", label_format)
    ws.write(row, 1, tab1_data.get("partName", ""), normal_border)
    row += 1
    ws.write(row, 0, "Jahresstückzahl", label_format)
    ws.write_number(row, 1, tab1_data.get("annualQty", 0), normal_border)
    row += 1
    ws.write(row, 0, "Losgröße", label_format)
    ws.write_number(row, 1, tab1_data.get("lotSize", 0), normal_border)
    row += 1
    ws.write(row, 0, "Ausschuss (%)", label_format)
    ws.write_number(row, 1, tab1_data.get("scrapPct", 0), normal_border)
    row += 1
    ws.write(row, 0, "SG&A (%)", label_format)
    ws.write_number(row, 1, tab1_data.get("sgaPct", 0), normal_border)
    row += 1
    ws.write(row, 0, "Profit (%)", label_format)
    ws.write_number(row, 1, tab1_data.get("profitPct", 0), normal_border)
    row += 1

    # Optional: Logo oben rechts
    logo_path = os.path.join(current_app.root_path, "static", "img", "logo.png")
    if os.path.exists(logo_path):
        ws.insert_image("I1", logo_path, {"x_scale": 0.6, "y_scale": 0.6})

    # ------------------------------------------------
    # 4) Material-Blöcke (Tab2) ab Zeile 10
    # ------------------------------------------------
    row = 10
    ws.write(row, 0, "Materialname", bold_border)
    ws.write(row, 1, "Materialmenge", header_format)
    ws.write(row, 2, "Materialpreis", header_format)
    ws.write(row, 5, "Materialkosten/100", header_format)
    ws.write(row, 6, "Material CO₂", header_format)

    row += 1
    matName = tab2_data.get("matName", "Aluminium")
    matWeight = tab2_data.get("matWeight", 0)
    matPrice = tab2_data.get("matPrice", 0)
    fremdVal = tab2_data.get("fremdValue", 0)
    mgk = tab2_data.get("matGK", 0)  # z. B. Material-Gemeinkosten

    # Füllbeispiel:
    ws.write(row, 0, matName, normal_border)
    ws.write_number(row, 1, matWeight, number_format)
    ws.write_number(row, 2, matPrice, number_format)

    # Tab4 könnte dir z. B. matEinzel=..., co2=... etc. geben
    # Machen wir hier als Demo:
    mat_cost_100 = tab4_summary.get("matEinzel", 0)
    mat_co2_100 = tab4_summary.get("matEinzel_co2", 0)
    ws.write_number(row, 5, mat_cost_100, number_format)
    ws.write_number(row, 6, mat_co2_100, number_format)

    row += 2
    ws.write(row, 0, "Fremdzukauf (€)", bold_border)
    ws.write_number(row, 5, fremdVal, number_format)
    # CO₂ z. B. 0 => anpassbar
    row += 2
    ws.write(row, 0, "MGK (%)", bold_border)
    ws.write_number(row, 5, mgk, number_format)
    # ...
    row += 1

    # Demo: Materialausschuss etc.
    ws.write(row, 0, "Materialausschuss", bold_border)
    mat_ausschuss_cost = tab4_summary.get("matAusschuss", 0)
    mat_ausschuss_co2 = tab4_summary.get("matAusschuss_co2", 0)
    ws.write_number(row, 5, mat_ausschuss_cost, number_format)
    ws.write_number(row, 6, mat_ausschuss_co2, number_format)
    row += 1

    ws.write(row, 0, "Materialkosten gesamt", bold_border)
    mat_ges_cost = tab4_summary.get("matGemein", 0)  # z. B. Summenkosten
    mat_ges_co2 = tab4_summary.get("matGemein_co2", 0)
    ws.write_number(row, 5, mat_ges_cost, number_format)
    ws.write_number(row, 6, mat_ges_co2, number_format)
    row += 2

    # ------------------------------------------------
    # 5) Fertigungsschritte ab z. B. Zeile 18
    # ------------------------------------------------
    ws.merge_range(row, 0, row, 9, "Fertigungsschritte", header_format)
    row += 1
    fert_headers = [
        "Arbeitsschritt", "Zyklus (s)", "MS (€/h)", "Lohn (€/h)",
        "Rüst (€/Los)", "Tooling/100", "FGK (%)", "CO₂ (kg/h)",
        "Kosten/100 (€)", "CO₂/100 (kg)"
    ]
    for col_idx, head in enumerate(fert_headers):
        ws.write(row, col_idx, head, header_format)
    row += 1
    start_fert_rows = row

    # bis zu 8 Steps
    for i, step in enumerate(tab3_steps[:8]):
        ws.write(row, 0, step.get("stepName", f"Step {i+1}"), normal_border)
        ws.write_number(row, 1, step.get("cycTime", 0), number_format)
        ws.write_number(row, 2, step.get("msRate", 0), number_format)
        ws.write_number(row, 3, step.get("lohnRate", 0), number_format)
        ws.write_number(row, 4, step.get("ruestVal", 0), number_format)
        ws.write_number(row, 5, step.get("tooling", 0), number_format)
        ws.write_number(row, 6, step.get("fgkPct", 0), number_format)
        ws.write_number(row, 7, step.get("co2Hour", 0), number_format)
        ws.write_number(row, 8, step.get("kosten_100", 0), number_format)
        ws.write_number(row, 9, step.get("co2_100", 0), number_format)
        row += 1

    # Summenzeile
    ws.write(row, 0, "Summe", bold_border)
    sum_cost = tab4_summary.get("summe_fert_cost", 0)
    sum_co2 = tab4_summary.get("summe_fert_co2", 0)
    ws.write_number(row, 8, sum_cost, number_format)
    ws.write_number(row, 9, sum_co2, number_format)
    row += 1

    # Optionale Zeilen: FGK, Tooling, Ausschuss, Herstellkosten ...
    ws.write(row, 0, "FGK", bold_border)
    fgk_cost = tab4_summary.get("fgk_cost", 0)
    fgk_co2 = tab4_summary.get("fgk_co2", 0)
    ws.write_number(row, 8, fgk_cost, number_format)
    ws.write_number(row, 9, fgk_co2, number_format)
    row += 1

    ws.write(row, 0, "Tooling", bold_border)
    tool_cost = tab4_summary.get("tool_cost", 0)
    tool_co2 = tab4_summary.get("tool_co2", 0)
    ws.write_number(row, 8, tool_cost, number_format)
    ws.write_number(row, 9, tool_co2, number_format)
    row += 1

    ws.write(row, 0, "Ausschuss", bold_border)
    ausschuss_cost = tab4_summary.get("fertAusschuss_cost", 0)
    ausschuss_co2 = tab4_summary.get("fertAusschuss_co2", 0)
    ws.write_number(row, 8, ausschuss_cost, number_format)
    ws.write_number(row, 9, ausschuss_co2, number_format)
    row += 1

    ws.write(row, 0, "Herstellkosten", bold_border)
    herstell_cost = tab4_summary.get("herstell", 0)
    herstell_co2 = tab4_summary.get("herstell_co2", 0)
    ws.write_number(row, 8, herstell_cost, number_format)
    ws.write_number(row, 9, herstell_co2, number_format)
    row += 2

    # SG&A, Profit, Gesamtkosten
    ws.write(row, 0, "SG&A", bold_border)
    ws.write_number(row, 1, tab4_summary.get("sga", 0), number_format)
    ws.write_number(row, 8, tab4_summary.get("sga_cost", 0), number_format)
    row += 1

    ws.write(row, 0, "Profit", bold_border)
    ws.write_number(row, 1, tab4_summary.get("profit", 0), number_format)
    ws.write_number(row, 8, tab4_summary.get("profit_cost", 0), number_format)
    row += 1

    ws.write(row, 0, "Gesamtkosten/100", bold_border)
    ws.write_number(row, 8, tab4_summary.get("total", 0), number_format)
    ws.write_number(row, 9, tab4_summary.get("co2_100", 0), number_format)
    row += 2

    # ------------------------------------------------
    # Fußzeile
    # ------------------------------------------------
    ws.write(row, 0, "Export-Datum:", label_format)
    ws.write(row, 1, datetime.now().strftime("%d.%m.%Y %H:%M"))

    # workbook schließen
    workbook.close()
    output.seek(0)
    return output, filename




